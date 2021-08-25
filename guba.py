import sys
import requests
import json
import urllib as UrlUtils
from bs4 import BeautifulSoup
import random
import ast
import ssl
ssl._create_default_https_context = ssl._create_unverified_context
import openpyxl


def request(num):
    # 设置浏览器头部信息
    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_2) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/47.0.2526.80 Safari/537.36'
    }
    url = 'https://guba.eastmoney.com/default,99_'+str(num)+'.html' # 需要请求的网页的链接
    html = requests.get(url,headers=headers)  # get方式请求数据
    # print(html.status_code)  # 查看请求的状态码（200表示请求正常）
    html.encoding = 'utf-8'  # 设置编码，防止由于编码问题导致文字错乱
    # print(html.text)  # 查看请求到的内容
    content = html.text
    return content




k=0
while(k<13):
    k = k + 1
    data = openpyxl.load_workbook('guba.xlsx')
    sheetnames = data.get_sheet_names()
    table = data.active
    nrows = table.max_row  # 获得行数

    con =request(k)
    soup = BeautifulSoup(con, "html.parser")
    # items 是一个 <listiterator object at 0x10a4b9950> 对象，不是一个list，但是可以循环遍历所有子节点。
    tmp = soup.find(attrs={'class': 'newlist'})
    items = tmp.find_all('li')
    projectList = []
    i = nrows+1
    for item in items:
            name = item.find_all('cite')[0].text
            table.cell(i, 1).value = name

            dd = item.find_all('cite')[1].text
            table.cell(i, 2).value = dd

            price = item.find_all('a')[0].text
            table.cell(i, 3).value = price

            score =item.find_all('a')[1].text
            table.cell(i, 4).value = score

            commentnum=item.find_all('cite')[2].text
            table.cell(i, 5).value = commentnum

            comment = item.find_all('cite')[3].text
            table.cell(i, 6).value = comment

            #comment = item.find_all('p')[0].text
            #table.cell(i, 6).value = comment

       #comment=item.find(attrs={'class': 'bottomcomment'}).string.strip()
       #table.cell(i, 6).value = comment
            print(k)
            #print(dd)
            #print(price)
            #print("aaaaaaaa")

            i = i+1
    data.save("guba.xlsx")

