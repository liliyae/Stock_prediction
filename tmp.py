import numpy as np
import openpyxl
import matplotlib.pyplot as plt
from sklearn.decomposition import PCA
from sklearn import preprocessing
import pandas as pd

def pca(X,k):#k is the contributing rate of principalcomponent you want
  #mean of each feature
  n_samples, n_features = X.shape
  mean=np.array([np.mean(X[:,i]) for i in range(n_features)])  #normalization
  norm_X=X-mean  #scatter matrix
  print(norm_X)
  scatter_matrix=np.dot(np.transpose(norm_X),norm_X)  #Calculate the eigenvectors and eigenvalues
  eig_val, eig_vec = np.linalg.eig(scatter_matrix)
  eig_pairs = [(np.abs(eig_val[i]), eig_vec[:,i]) for i in range(n_features)]  # sort eig_vec based on eig_val from highest to lowest
  eig_pairs.sort(reverse=True)  # select the top k eig_vec
  rat = 0
  num = 0
  while(rat<=k):
    rat =rat+eig_val[num]
    num=num+1

  feature=np.array([ele[1] for ele in eig_pairs[:num]])  #get new data
  data=np.dot(norm_X,np.transpose(feature))
  return data

data = openpyxl.load_workbook('data.xlsx')
sheetnames = data.get_sheet_names()
table = data.active
nrows = table.max_row  # 获得行数
cols= table.max_column

x = np.empty([1326,66], dtype = float)

for i in range(3,1329):
  for j in range(2,68):
    x[i-3,j-2]=table.cell(i,j).value

tmp = pca(x,80)
print(tmp)
print(tmp.shape)



