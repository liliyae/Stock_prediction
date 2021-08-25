import urllib
import os
import time
import re
import logging
import requests
import urllib3
# json解析库,对应到lxml
import json
# 全局取消证书验证
import io
import sys
# 改变标准输出的默认编码
# utf-8中文乱码 有些表情print不进去
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='gb18030')
import random
import ast
import ssl
ssl._create_default_https_context = ssl._create_unverified_context
import openpyxl


data = openpyxl.load_workbook('data.xlsx')
sheetnames = data.get_sheet_names()
table = data.active
nrows = table.max_row  # 获得行数

path="股票数据.xlsx"
p = openpyxl.load_workbook(path)
sheetnames = p.get_sheet_names()
ptable = p.active
prows = ptable.max_row  # 获得行数

i=3
j=3
while(i<=nrows):
    tmp = ptable.cell(j, 1).value
    day = table.cell(i,1).value
    if(day<=tmp):
        j=j+1
        tmp = ptable.cell(j, 1).value

    for index in range(2,8):
        print(j)
        table.cell(i,index+28).value=ptable.cell(j-1, index).value
    i=i+1

data.save("data.xlsx")
p.save(path)





