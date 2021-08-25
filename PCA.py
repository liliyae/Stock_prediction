import numpy as np
import openpyxl
import matplotlib.pyplot as plt
from sklearn.decomposition import PCA
from sklearn import preprocessing
import pandas as pd


data = openpyxl.load_workbook('data.xlsx')
sheetnames = data.get_sheet_names()
table = data.active
nrows = table.max_row  # 获得行数
cols= table.max_column

x = np.empty([1326,66], dtype = float)

for i in range(3,1329):
  for j in range(2,68):
    x[i-3,j-2]=table.cell(i,j).value

min_max_scaler = preprocessing.MinMaxScaler()
x_minMax = min_max_scaler.fit_transform(x)#归一化

#mean of each feature
n_samples, n_features = x_minMax.shape
mean=np.array([np.mean(x_minMax[:,i]) for i in range(n_features)])  #normalization
xd = x_minMax-mean  #去中心化
print(xd)


model=PCA(n_components=0.95)
#model=PCA(n_components=0.98)
model.fit(xd)

X_new=model.fit_transform(xd)
Maxcomponent=model.components_
ratio=model.explained_variance_ratio_
score=model.score(xd)
print('降维后的数据:',X_new)
print('返回具有最大方差的成分:',Maxcomponent)
print('保留主成分的方差贡献率:',ratio)
print('所有样本的log似然平均值:',score)
print('奇异值:',model.singular_values_)
print('噪声协方差:',model.noise_variance_)

g1=plt.figure(1,figsize=(8,6))
plt.scatter(X_new[:,0],X_new[:,1],c='r',cmap=plt.cm.Set1, edgecolor='k', s=40)
plt.xlabel('x1')
plt.ylabel('x2')
plt.title('After the dimension reduction')
plt.show()

