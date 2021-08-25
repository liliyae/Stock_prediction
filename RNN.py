import openpyxl
from sklearn.decomposition import PCA
from sklearn import preprocessing
import numpy as np
import tensorflow as tf
from tensorflow.keras.layers import Dropout, Dense, SimpleRNN
import matplotlib.pyplot as plt
import os
#import pandas as pd
from sklearn.preprocessing import MinMaxScaler
from sklearn.metrics import mean_squared_error, mean_absolute_error
import math

#输入数据
data = openpyxl.load_workbook('data.xlsx')
sheetnames = data.get_sheet_names()
table = data.active
nrows = table.max_row  # 获得行数
cols= table.max_column

x = np.empty([1351,66], dtype = float)

for i in range(3,1354):
  for j in range(2,68):
    x[i-3,j-2]=table.cell(i,j).value

#输出数据####################
out = openpyxl.load_workbook('out.xlsx')
sheet = out.get_sheet_names()
outtable = out.active

outdata = np.empty([1326,1], dtype = float)
for i in range(3,1329):
    outdata[i - 3] = outtable.cell(i, 2).value


##########################


min_max_scaler = preprocessing.MinMaxScaler()
x_minMax = min_max_scaler.fit_transform(x)#归一化
#outdata = min_max_scaler.fit_transform(outdata)#归一化

#mean of each feature
n_samples, n_features = x_minMax.shape
mean=np.array([np.mean(x_minMax[:,i]) for i in range(n_features)])  #normalization
xd = x_minMax-mean  #去中心化
print(xd)


model=PCA(n_components=0.95)
#model=PCA(n_components=0.98)
model.fit(xd)

X_new=model.fit_transform(xd)
Maxcomponent=model.components_
ratio=model.explained_variance_ratio_
score=model.score(xd)
print('降维后的数据:',X_new)
print('返回具有最大方差的成分:',Maxcomponent)
print('保留主成分的方差贡献率:',ratio)
print('所有样本的log似然平均值:',score)
print('奇异值:',model.singular_values_)
print('噪声协方差:',model.noise_variance_)

days = 20

training_set = X_new[0:1326]  # 前1000天的开盘价作为训练集,表格从0开始计数，2:3 是提取[2:3)列，前闭后开,故提取出C列开盘价
test_set = X_new[0:1326] # 后300天的开盘价作为测试集
training_out = outdata[0:1326]
test_out = outdata[0:1326]
ptest = X_new[1327-days:1351]

# 归一化
#sc = MinMaxScaler(feature_range=(0, 1))  # 定义归一化：归一化到(0，1)之间
#training_set_scaled = sc.fit_transform(training_set)  # 求得训练集的最大值，最小值这些训练集固有的属性，并在训练集上进行归一化
#test_set = sc.transform(test_set)  # 利用训练集的属性对测试集进行归一化


x_train = []
y_train = []

x_test = []
y_test = []

pptest=[]

# 测试集：csv表格中前2426-300=2126天数据
# 利用for循环，遍历整个训练集，提取训练集中连续60天的开盘价作为输入特征x_train，第61天的数据作为标签，for循环共构建2426-300-60=2066组数据。


for i in range(days, len(training_set)):
    x_train.append(training_set[i - days:i, 0:5])
    y_train.append(training_out[i])

# 对训练集进行打乱
np.random.seed(7)
np.random.shuffle(x_train)
np.random.seed(7)
np.random.shuffle(y_train)
tf.random.set_seed(7)
# 将训练集由list格式变为array格式
x_train, y_train = np.array(x_train), np.array(y_train)

# 使x_train符合RNN输入要求：[送入样本数， 循环核时间展开步数， 每个时间步输入特征个数]。
# 此处整个数据集送入，送入样本数为x_train.shape[0]即2066组数据；输入60个开盘价，预测出第61天的开盘价，循环核时间展开步数为60; 每个时间步送入的特征是某一天的开盘价，只有1个数据，故每个时间步输入特征个数为1
x_train = np.reshape(x_train, (x_train.shape[0], days, 5))
# 测试集：csv表格中后300天数据
# 利用for循环，遍历整个测试集，提取测试集中连续60天的开盘价作为输入特征x_train，第61天的数据作为标签，for循环共构建300-60=240组数据。
for i in range(days, len(test_set)):
    x_test.append(test_set[i - days:i, 0:5])
    y_test.append(test_out[i])

for i in range(days, len(ptest)):
    pptest.append(ptest[i - days:i, 0:5])
pptest= np.array(pptest)
pptest = np.reshape(pptest, (pptest.shape[0], days, 5))

# 测试集变array并reshape为符合RNN输入要求：[送入样本数， 循环核时间展开步数， 每个时间步输入特征个数]
x_test, y_test = np.array(x_test), np.array(y_test)
x_test = np.reshape(x_test, (x_test.shape[0], days, 5))

model = tf.keras.Sequential([
    SimpleRNN(80, return_sequences=True),
    Dropout(0.2),
    SimpleRNN(100),
    Dropout(0.2),
    Dense(1)
])

model.compile(optimizer=tf.keras.optimizers.Adam(0.001),
              loss='mean_squared_error')  # 损失函数用均方误差
# 该应用只观测loss数值，不观测准确率，所以删去metrics选项，一会在每个epoch迭代显示时只显示loss值

checkpoint_save_path = "./checkpoint/rnn_stock.ckpt"

if os.path.exists(checkpoint_save_path + '.index'):
    print('-------------load the model-----------------')
    model.load_weights(checkpoint_save_path)

cp_callback = tf.keras.callbacks.ModelCheckpoint(filepath=checkpoint_save_path,
                                                 save_weights_only=True,
                                                 save_best_only=True,
                                                 monitor='val_loss')

history = model.fit(x_train, y_train, batch_size=64, epochs=50, validation_data=(x_test, y_test), validation_freq=1,
                    callbacks=[cp_callback])

model.summary()

file = open('./weights.txt', 'w')  # 参数提取
for v in model.trainable_variables:
    file.write(str(v.name) + '\n')
    file.write(str(v.shape) + '\n')
    file.write(str(v.numpy()) + '\n')
file.close()

loss = history.history['loss']
val_loss = history.history['val_loss']


plt.plot(loss, label='Training Loss')
plt.plot(val_loss, label='Validation Loss')
plt.title('Training and Validation Loss')
plt.legend()
plt.show()

################## predict ######################
# 测试集输入模型进行预测
predicted_stock_price = model.predict(pptest)
# 对预测数据还原---从（0，1）反归一化到原始范围
#predicted_stock_price = sc.inverse_transform(predicted_stock_price)
# 对真实数据还原---从（0，1）反归一化到原始范围
#real_stock_price = sc.inverse_transform(test_set[60:])
real_stock_price = y_test[0:]


# 画出真实数据和预测数据的对比曲线
fig = plt.figure(figsize=(12, 10))
#plt.plot(real_stock_price, color='red', label='MaoTai Stock Price')
plt.plot(predicted_stock_price, color='blue', label='Predicted MaoTai Stock Price')
plt.title('MaoTai Stock Price Prediction')
plt.xlabel('Time')
plt.ylabel('MaoTai Stock Price')
plt.legend()
plt.show()
plt.savefig('D:\\光华\\codes\\plot2.png')

##########evaluate##############
# calculate MSE 均方误差 ---> E[(预测值-真实值)^2] (预测值减真实值求平方后求均值)
mse = mean_squared_error(predicted_stock_price, real_stock_price)
# calculate RMSE 均方根误差--->sqrt[MSE]    (对均方误差开方)
rmse = math.sqrt(mean_squared_error(predicted_stock_price, real_stock_price))
# calculate MAE 平均绝对误差----->E[|预测值-真实值|](预测值减真实值求绝对值后求均值）
mae = mean_absolute_error(predicted_stock_price, real_stock_price)
print('均方误差: %.6f' % mse)
print('均方根误差: %.6f' % rmse)
print('平均绝对误差: %.6f' % mae)